"""Write curated events to a persistent .xlsx spreadsheet and run_log.md.

Key behaviours (Tasks 8, 9, 13, 14):
- Output is ``agent_research.xlsx`` — the source of truth for all events.
- Columns: Event, Venue, Location, Date, URL, Sources, Poster URL, Summary, Added.
- The spreadsheet **accumulates**: existing rows are merged, past events removed,
  new events added.  Nothing is overwritten by new LLM text.
- Semantic deduplication (Task 13): before adding a row the pipeline checks
  whether an event with the same (act, date) already exists (venue ignored for
  exact matches because venue text can vary slightly).  If so the new URL is
  appended to that row's Sources column (when the domain differs) rather than
  creating a duplicate row.
- Partial-name deduplication (Task 14): when act names only partially match
  (e.g. "Singer 1" vs "Singer 1, with Singer 2") but the venue AND date are
  identical, the two entries are treated as the same event.  The longer act
  name is kept as the canonical name; the duplicate URL is added to Sources if
  from a different domain.
- ``run_log.md`` is unchanged.
"""

from __future__ import annotations

import logging
import os
from datetime import date, datetime
from pathlib import Path
from urllib.parse import urlparse

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from agent import config
from agent.display_time import format_generated_timestamp
from agent.event_window import (
    parse_event_sort_date,
    split_title_parts,
    utc_today,
)
from agent.models import Resource

logger = logging.getLogger(__name__)

RESEARCH_FILENAME = "agent_research.xlsx"
RUN_LOG_FILENAME = "run_log.md"

# ── Schema ────────────────────────────────────────────────────────────────────
# Column names in display order.  The Date cell stores a Python date object
# so Excel can sort natively; the number format renders it as "Wed 7 May 2026".
# Sources stores newline-separated alternative URLs for the same gig.
_COLS = [
    "Event", "Venue", "Location", "Date", "URL",
    "Sources", "Poster URL", "Summary", "Added",
]
_DATE_FORMAT = "ddd d mmm yyyy"

# 0-based indices into _COLS
_IDX_EVENT    = 0
_IDX_VENUE    = 1
_IDX_LOCATION = 2
_IDX_DATE     = 3
_IDX_URL      = 4
_IDX_SOURCES  = 5
_IDX_POSTER   = 6
_IDX_SUMMARY  = 7
_IDX_ADDED    = 8

# Column widths (0 = hide the column)
_COL_WIDTHS = [30, 22, 16, 18, 55, 45, 0, 40, 20]


def output_directory() -> Path:
    """Repo ``data/`` by default; override with OUTPUT_DIR or AGENT_AI_DIR in ``.env``."""
    return config.OUTPUT_DIR


# ── Domain / dedup helpers ────────────────────────────────────────────────────


def _domain(url: str) -> str:
    """Return the bare domain of a URL (lowercase, no 'www.')."""
    try:
        netloc = urlparse(url).netloc.lower()
        return netloc.lstrip("www.") if netloc.startswith("www.") else netloc
    except Exception:
        return ""


def _dedup_key_from_row(row: list) -> tuple[str, date] | None:
    """(act_lower, date) used to detect semantic duplicates.

    Venue is intentionally excluded because venue text can vary slightly
    across sources while still referring to the same event.
    Returns None when the row lacks enough data to compare.
    """
    act = str(row[_IDX_EVENT] or "").strip().lower()
    d   = _row_date(row)
    if not act or d is None:
        return None
    return (act, d)


def _dedup_key_from_resource(r: Resource) -> tuple[str, date] | None:
    """Same key shape as ``_dedup_key_from_row``, but derived from a Resource."""
    act, _venue, _location = split_title_parts(r.title or "")
    d = parse_event_sort_date(r.date)
    if not act.strip() or d is None:
        return None
    return (act.strip().lower(), d)


def _add_source(row: list, url: str) -> bool:
    """Append *url* to a row's Sources cell when it is from a different domain.

    Returns True when a new source was actually added.
    """
    primary_url = str(row[_IDX_URL] or "").strip()
    new_domain  = _domain(url)
    if not new_domain or new_domain == _domain(primary_url):
        return False  # same domain — not worth recording as an alternative

    current = str(row[_IDX_SOURCES] or "").strip()
    existing_urls = [u.strip() for u in current.split("\n") if u.strip()]
    if url in existing_urls:
        return False

    existing_urls.append(url)
    row[_IDX_SOURCES] = "\n".join(existing_urls)
    return True


def _acts_partially_match(act1: str, act2: str) -> bool:
    """True when one act name is a substring of the other (case-insensitive).

    Requires both names to be at least 4 characters so trivially short strings
    (e.g. "The") don't produce false positives.  This catches common cases like
    "Singer 1" vs "Singer 1, with Singer 2" where one name simply adds support
    act information to the headline act.
    """
    a = act1.strip().lower()
    b = act2.strip().lower()
    if len(a) < 4 or len(b) < 4:
        return False
    return a in b or b in a


def _find_partial_act_match(
    new_act: str,
    new_date: date,
    new_venue: str,
    existing: dict[str, list],
) -> str | None:
    """Scan existing rows for a partial act-name match on the same venue + date.

    Venue is used as a required tie-breaker here (unlike exact-name dedup) to
    prevent merging genuinely different acts that happen to share a name fragment
    on different stages.

    Returns the url_key of the first matching row, or None.
    """
    new_venue_norm = new_venue.strip().lower()
    for url_key, row in existing.items():
        if _row_date(row) != new_date:
            continue
        if str(row[_IDX_VENUE] or "").strip().lower() != new_venue_norm:
            continue
        if _acts_partially_match(new_act, str(row[_IDX_EVENT] or "")):
            return url_key
    return None


# ── Row ↔ Resource conversion ─────────────────────────────────────────────────


def _resource_to_row(r: Resource) -> list:
    """Convert a Resource to a list matching ``_COLS`` order."""
    act, venue, location = split_title_parts(r.title or "")
    raw_date = parse_event_sort_date(r.date)
    return [
        act or "—",                           # Event
        venue,                                # Venue
        location,                             # Location
        raw_date,                             # Date (Excel date cell)
        (r.url or "").strip(),                # URL
        "",                                   # Sources (populated by dedup logic)
        (r.thumbnail_url or "").strip(),      # Poster URL (hidden)
        (r.summary or "").strip(),            # Summary
        format_generated_timestamp(),         # Added
    ]


def _row_date(row: list) -> date | None:
    """Extract the date from a row; normalise datetime → date."""
    val = row[_IDX_DATE] if len(row) > _IDX_DATE else None
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    return None


def _row_to_resource(row: list) -> Resource:
    """Reconstruct a Resource from a spreadsheet row.

    The title is rebuilt in "Act @ Venue, Location" format so that
    ``split_title_parts`` in html_output / notion_output can decompose it
    the same way it was originally split.
    """
    act      = str(row[_IDX_EVENT]    or "").strip()
    venue    = str(row[_IDX_VENUE]    or "").strip()
    location = str(row[_IDX_LOCATION] or "").strip()
    url      = str(row[_IDX_URL]      or "").strip()
    poster   = str(row[_IDX_POSTER]   or "").strip()
    summary  = str(row[_IDX_SUMMARY]  or "").strip()
    raw_date = _row_date(row)

    venue_loc = ", ".join(filter(None, [venue, location]))
    title = f"{act} @ {venue_loc}" if venue_loc else act

    return Resource(
        title=title,
        url=url,
        date=raw_date.isoformat() if raw_date else "",
        thumbnail_url=poster or None,
        summary=summary,
    )


def load_spreadsheet_resources(path: Path | None = None) -> list[Resource]:
    """Read the current spreadsheet and return its rows as Resource objects.

    This is the public API for getting the source-of-truth event list.
    """
    if path is None:
        path = output_directory() / RESEARCH_FILENAME
    rows = _load_existing_rows(path)
    return [_row_to_resource(row) for row in rows.values()]


# ── Spreadsheet I/O ───────────────────────────────────────────────────────────


def _load_existing_rows(path: Path) -> dict[str, list]:
    """Load the spreadsheet as an ordered dict keyed by URL (lowercase).

    The loader is lenient about the Sources column: if every other expected
    column is present but Sources is absent (i.e. a spreadsheet written before
    Task 13 was deployed), the column is silently synthesised as empty.  Any
    other schema mismatch logs a warning and returns an empty dict.
    """
    rows: dict[str, list] = {}
    if not path.exists():
        return rows
    try:
        wb = load_workbook(path)
        ws = wb.active
        header = [str(c.value or "") for c in next(ws.iter_rows(min_row=1, max_row=1))]

        missing = [n for n in _COLS if n not in header]
        if missing and missing != ["Sources"]:
            # Too many missing columns — schema is incompatible.
            logger.warning(
                "Spreadsheet is missing unexpected columns %s; "
                "ignoring existing file (header: %s).",
                missing,
                header,
            )
            return rows

        # Build column index; Sources gets -1 when absent (pre-Task-13 file).
        col_index = {
            name: (header.index(name) if name in header else -1)
            for name in _COLS
        }
        url_col = col_index["URL"]

        for raw_row in ws.iter_rows(min_row=2, values_only=True):
            if url_col < 0 or url_col >= len(raw_row):
                continue
            url_key = str(raw_row[url_col] or "").strip().lower()
            if not url_key.startswith("http"):
                continue
            row_list = [
                (raw_row[col_index[c]] if col_index[c] >= 0 and col_index[c] < len(raw_row) else None)
                for c in _COLS
            ]
            # Ensure Sources is a string (not None) for safe appending later.
            if row_list[_IDX_SOURCES] is None:
                row_list[_IDX_SOURCES] = ""
            rows[url_key] = row_list
    except Exception as exc:
        logger.warning("Could not read existing spreadsheet (%s); starting fresh.", exc)
    return rows


def _write_workbook(path: Path, rows: dict[str, list]) -> None:
    """Sort by date and save all rows to the spreadsheet file."""
    def sort_key(row: list) -> date:
        d = _row_date(row)
        return d if d is not None else date.max

    sorted_rows = sorted(rows.values(), key=sort_key)

    wb = Workbook()
    ws = wb.active
    ws.title = "Events"

    # Header row
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="2E4057")
    for col_i, name in enumerate(_COLS, start=1):
        cell = ws.cell(row=1, column=col_i, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Data rows
    for row_i, row_vals in enumerate(sorted_rows, start=2):
        for col_i, value in enumerate(row_vals[: len(_COLS)], start=1):
            cell = ws.cell(row=row_i, column=col_i, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if col_i == _IDX_DATE + 1 and isinstance(value, (date, datetime)):
                cell.number_format = _DATE_FORMAT
            elif col_i == _IDX_URL + 1 and str(value or "").startswith("http"):
                cell.hyperlink = str(value)
                cell.font = Font(color="0563C1", underline="single")

    # Column widths / visibility
    for col_i, w in enumerate(_COL_WIDTHS, start=1):
        letter = get_column_letter(col_i)
        if w == 0:
            ws.column_dimensions[letter].hidden = True
        else:
            ws.column_dimensions[letter].width = w

    ws.freeze_panes = "A2"
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = path.with_name(path.name + ".tmp")
    wb.save(tmp_path)
    try:
        os.replace(tmp_path, path)
    except PermissionError:
        fallback = path.with_name(path.stem + "_unlock_me" + path.suffix)
        try:
            if fallback.exists():
                fallback.unlink()
        except OSError:
            pass
        os.replace(tmp_path, fallback)
        logger.error(
            "%s is locked (usually open in Excel). Latest data saved as %s — "
            "close the workbook and rename it, or re-run the agent.",
            path.name, fallback.name,
        )
        raise PermissionError(
            f"{path.name} is in use (close it in Excel). "
            f"Latest data saved as {fallback.name}."
        ) from None
    logger.info("Spreadsheet written: %d events → %s", len(sorted_rows), path)


# ── Merge + expire + dedup ────────────────────────────────────────────────────


def merge_and_write(new_resources: list[Resource]) -> tuple[int, int, int]:
    """Merge new events into the persistent spreadsheet.

    Steps:
    1. Load existing rows from disk.
    2. Drop rows whose event date is in the past.
    3. For each new resource:
       a. Exact URL match → skip.
       b. Same act name + date (venue ignored) → exact semantic duplicate:
          add URL to Sources when domain differs.
       c. One act name is a substring of the other + same venue + same date →
          partial-name duplicate: keep the longer act name; add URL to Sources.
       d. Otherwise → insert as a new row.
    4. Sort soonest-first and save.

    Returns:
        (added, skipped_duplicate, removed_past) counts.
    """
    out_dir = output_directory()
    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / RESEARCH_FILENAME
    today = utc_today()

    existing = _load_existing_rows(path)

    # Remove past events
    before = len(existing)
    existing = {k: v for k, v in existing.items() if (_row_date(v) or date.max) >= today}
    removed_past = before - len(existing)

    # Build semantic dedup index: (act_lower, venue_lower, date) → url_key
    dedup_index: dict[tuple, str] = {}
    for url_key, row in existing.items():
        dk = _dedup_key_from_row(row)
        if dk:
            dedup_index[dk] = url_key

    added = 0
    skipped = 0
    for r in new_resources:
        url = (r.url or "").strip()
        url_key = url.lower()
        if not url_key.startswith("http"):
            continue

        # Decompose title once — used by several checks below.
        act, venue, _location = split_title_parts(r.title or "")
        r_date = parse_event_sort_date(r.date)

        # (a) Exact URL duplicate
        if url_key in existing:
            skipped += 1
            continue

        # (b) Exact semantic duplicate — same act name + date (venue ignored).
        dk = _dedup_key_from_resource(r)
        if dk and dk in dedup_index:
            match_key = dedup_index[dk]
            source_added = _add_source(existing[match_key], url)
            if source_added:
                logger.debug(
                    "Exact-name duplicate for '%s' — added %s to Sources.",
                    r.title, url,
                )
            skipped += 1
            continue

        # (c) Partial-name duplicate — one act name contains the other, AND the
        #     venue + date are identical.  Keep the longer (more informative)
        #     act name as the canonical one.
        if act and r_date:
            partial_key = _find_partial_act_match(act, r_date, venue, existing)
            if partial_key:
                existing_act = str(existing[partial_key][_IDX_EVENT] or "")
                if len(act) > len(existing_act):
                    existing[partial_key][_IDX_EVENT] = act
                    logger.debug(
                        "Partial-name duplicate: upgraded canonical name '%s' → '%s'.",
                        existing_act, act,
                    )
                _add_source(existing[partial_key], url)
                skipped += 1
                continue

        # (d) Genuinely new event
        new_row = _resource_to_row(r)
        existing[url_key] = new_row
        if dk:
            dedup_index[dk] = url_key
        added += 1

    _write_workbook(path, existing)
    logger.info(
        "Spreadsheet: +%d added, %d duplicate/skipped, %d expired removed → %d total",
        added, skipped, removed_past, len(existing),
    )
    return added, skipped, removed_past


# ── Public API (called from graph_nodes) ─────────────────────────────────────


def write_output(
    resources: list[Resource],
    *,
    append_log_only: bool,
    log_line: str,
) -> None:
    """Merge new events into the spreadsheet and append to run_log.md."""
    out_dir = output_directory()
    out_dir.mkdir(parents=True, exist_ok=True)
    log_path = out_dir / RUN_LOG_FILENAME

    merge_and_write(resources)
    _append_log(log_path, log_line)
    logger.info("Appended run log entry to %s", log_path)


def _append_log(log_path: Path, log_line: str) -> None:
    """Append a single log line to the run log, creating it when missing."""
    block = f"\n- {log_line}\n"
    existing = log_path.read_text(encoding="utf-8") if log_path.exists() else ""
    if not existing.strip():
        log_path.write_text("# Run log\n\n" + block.lstrip(), encoding="utf-8")
    else:
        log_path.write_text(existing.rstrip() + block, encoding="utf-8")
