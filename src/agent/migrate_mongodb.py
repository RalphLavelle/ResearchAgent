"""One-time migration from file-based storage to MongoDB."""

from __future__ import annotations

import json
import logging
import shutil
from pathlib import Path
from uuid import uuid4

from openpyxl import load_workbook

from agent import config, image_store
from agent.event_store import save_existing_rows
from agent.image_cache import cache_thumbnails
from agent.local_output import (
    RESEARCH_FILENAME,
    _COLS,
    _resource_to_row,
    _row_to_resource,
)
from agent.topics import TopicEntry, load_topics, topic_data_dir

logger = logging.getLogger(__name__)

_INDEX_FILENAME = "_index.json"
_IMAGES_SUBDIR = "images"
_KNOWN_EXTS = (".jpg", ".jpeg", ".png", ".webp", ".gif", ".avif")


def _load_rows_from_xlsx(path: Path) -> dict[str, list]:
    """Read legacy spreadsheet into row dict."""
    # Temporarily read via openpyxl inline (migration-only).
    rows: dict[str, list] = {}
    if not path.exists():
        return rows
    wb = load_workbook(path)
    ws = wb.active
    header = [str(c.value or "") for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col_index = {name: header.index(name) if name in header else -1 for name in _COLS}
    url_col = col_index["URL"]
    for raw_row in ws.iter_rows(min_row=2, values_only=True):
        if url_col < 0 or url_col >= len(raw_row):
            continue
        url_key = str(raw_row[url_col] or "").strip().lower()
        if not url_key.startswith("http"):
            continue
        row_list = [
            (
                raw_row[col_index[c]]
                if col_index[c] >= 0 and col_index[c] < len(raw_row)
                else None
            )
            for c in _COLS
        ]
        if row_list[5] is None:
            row_list[5] = ""
        sid = str(row_list[9] or "").strip() or str(uuid4())
        while sid in rows:
            sid = str(uuid4())
        row_list[9] = sid
        rows[sid] = row_list
    return rows


def _resources_from_events_json(path: Path) -> list[Resource]:
    if not path.exists():
        return []
    data = json.loads(path.read_text(encoding="utf-8"))
    out: list[Resource] = []
    for ev in data.get("events") or []:
        if not isinstance(ev, dict):
            continue
        name = str(ev.get("eventName") or "").strip()
        venue = str(ev.get("venue") or "").strip()
        title = f"{name} @ {venue}" if venue else name
        thumb = ev.get("thumbnailUrl")
        out.append(
            Resource(
                id=str(ev.get("id") or uuid4()),
                title=title,
                url=str(ev.get("url") or ""),
                date=str(ev.get("date") or ""),
                summary=str(ev.get("summary") or ""),
                thumbnail_url=str(thumb).strip() if thumb else None,
            )
        )
    return out


def _guess_content_type(path: Path) -> str:
    ext = path.suffix.lower()
    mapping = {
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".png": "image/png",
        ".webp": "image/webp",
        ".gif": "image/gif",
        ".avif": "image/avif",
    }
    return mapping.get(ext, "application/octet-stream")


def _migrate_images_from_folder(
    images_dir: Path,
    db_name: str,
    *,
    index: dict,
) -> dict[str, str]:
    """Upload on-disk poster files; return event_id → image_id."""
    event_to_image: dict[str, str] = {}
    if not images_dir.is_dir():
        return event_to_image

    events_index = (index or {}).get("events") or {}
    for eid, rec in events_index.items():
        if not isinstance(rec, dict):
            continue
        fname = str(rec.get("file") or "").strip()
        source = str(rec.get("source") or "").strip()
        if not fname:
            continue
        fpath = images_dir / fname
        if not fpath.is_file():
            continue
        data = fpath.read_bytes()
        ctype = _guess_content_type(fpath)
        image_store.store_image(
            db_name,
            image_id=fname,
            source_url=source or f"file://{fname}",
            data=data,
            content_type=ctype,
        )
        event_to_image[str(eid)] = fname

    # Orphan files without index — store by filename as id.
    for path in images_dir.iterdir():
        if not path.is_file() or path.suffix.lower() not in _KNOWN_EXTS:
            continue
        if path.name == _INDEX_FILENAME:
            continue
        if path.name in event_to_image.values():
            continue
        data = path.read_bytes()
        ctype = _guess_content_type(path)
        image_store.store_image(
            db_name,
            image_id=path.name,
            source_url=f"file://{path.name}",
            data=data,
            content_type=ctype,
        )
    return event_to_image


def migrate_topic(
    topic_id: str,
    entry: TopicEntry,
    *,
    data_base: Path,
    remove_files: bool = True,
) -> dict[str, int]:
    """Migrate one topic's legacy files into MongoDB."""
    folder = topic_data_dir(data_base, topic_id)
    db_name = entry.db
    stats = {"events": 0, "images": 0, "files_removed": 0}

    xlsx = folder / RESEARCH_FILENAME
    events_json = folder / "events.json"
    images_dir = folder / _IMAGES_SUBDIR

    rows: dict[str, list] = {}
    if xlsx.exists():
        rows = _load_rows_from_xlsx(xlsx)
    elif events_json.exists():
        for r in _resources_from_events_json(events_json):
            rows[r.id] = _resource_to_row(r)

    if rows:
        save_existing_rows(db_name, rows)
        stats["events"] = len(rows)
    index: dict = {}
    index_path = images_dir / _INDEX_FILENAME
    if index_path.exists():
        try:
            index = json.loads(index_path.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            index = {}

    event_images = _migrate_images_from_folder(images_dir, db_name, index=index)
    stats["images"] = len(image_store.list_image_ids(db_name))

    if event_images:
        image_store.bulk_update_event_image_ids(
            db_name,
            {eid: iid for eid, iid in event_images.items()},
        )

    if remove_files:
        for name in (RESEARCH_FILENAME, "events.json"):
            p = folder / name
            if p.exists():
                p.unlink()
                stats["files_removed"] += 1
        if images_dir.is_dir():
            shutil.rmtree(images_dir)
            stats["files_removed"] += 1

    logger.info(
        "Migrated topic %s → db=%s (%d events, %d images)",
        topic_id,
        db_name,
        stats["events"],
        stats["images"],
    )
    return stats


def migrate_all_topics(*, remove_files: bool = True) -> dict[str, dict[str, int]]:
    """Migrate every registered topic from ``data/<topic_id>/`` to MongoDB."""
    reg = load_topics(config.TOPICS_CONFIG_PATH)
    results: dict[str, dict[str, int]] = {}
    for topic_id, entry in reg.topics.items():
        results[topic_id] = migrate_topic(
            topic_id,
            entry,
            data_base=config.DATA_BASE_DIR,
            remove_files=remove_files,
        )
    return results
