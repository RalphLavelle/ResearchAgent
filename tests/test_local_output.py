"""Tests for the accumulating spreadsheet output (Tasks 8, 9, 13)."""

from datetime import date, timedelta
from pathlib import Path

import pytest
from openpyxl import load_workbook

from agent.local_output import (
    RESEARCH_FILENAME,
    _IDX_SOURCES,
    _IDX_URL,
    load_spreadsheet_resources,
    merge_and_write,
)
from agent.models import Resource


def _make_resource(title: str, url: str, days_ahead: int = 5) -> Resource:
    d = (date.today() + timedelta(days=days_ahead)).isoformat()
    return Resource(title=title, url=url, date=d)


def _urls_from_sheet(path: Path) -> list[str]:
    wb = load_workbook(path)
    ws = wb.active
    return [str(row[_IDX_URL] or "") for row in ws.iter_rows(min_row=2, values_only=True)]


def _sources_from_sheet(path: Path) -> list[str]:
    wb = load_workbook(path)
    ws = wb.active
    return [str(row[_IDX_SOURCES] or "") for row in ws.iter_rows(min_row=2, values_only=True)]


# ── Basic write / read ────────────────────────────────────────────────────────

def test_merge_creates_spreadsheet(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr("agent.config.OUTPUT_DIR", tmp_path)
    r = _make_resource("Band A @ Venue X, Brisbane", "https://example.com/a")
    added, skipped, removed = merge_and_write([r])
    assert added == 1
    assert skipped == 0
    assert removed == 0
    assert (tmp_path / RESEARCH_FILENAME).exists()


def test_title_splits_into_columns(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr("agent.config.OUTPUT_DIR", tmp_path)
    r = _make_resource("The Beths @ The Tivoli, Brisbane", "https://example.com/beths")
    merge_and_write([r])
    wb = load_workbook(tmp_path / RESEARCH_FILENAME)
    ws = wb.active
    row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    assert row[0] == "The Beths"      # Event
    assert row[1] == "The Tivoli"     # Venue
    assert row[2] == "Brisbane"       # Location


def test_new_url_added_on_second_run(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr("agent.config.OUTPUT_DIR", tmp_path)
    r1 = _make_resource("Band A", "https://example.com/a")
    r2 = _make_resource("Band B", "https://example.com/b")
    merge_and_write([r1])
    added, _, _ = merge_and_write([r2])
    assert added == 1
    urls = _urls_from_sheet(tmp_path / RESEARCH_FILENAME)
    assert "https://example.com/a" in urls
    assert "https://example.com/b" in urls


def test_past_events_removed(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr("agent.config.OUTPUT_DIR", tmp_path)
    past = Resource(title="Old Gig", url="https://example.com/old", date="2020-01-01")
    future = _make_resource("New Gig", "https://example.com/new")
    merge_and_write([past, future])
    _, _, removed = merge_and_write([])
    assert removed == 1
    urls = _urls_from_sheet(tmp_path / RESEARCH_FILENAME)
    assert "https://example.com/old" not in urls
    assert "https://example.com/new" in urls


def test_load_spreadsheet_resources_roundtrip(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr("agent.config.OUTPUT_DIR", tmp_path)
    r1 = _make_resource("The Beths @ The Tivoli, Brisbane", "https://example.com/beths")
    r2 = _make_resource("Open Mic Night @ Burleigh Bazaar, Gold Coast", "https://example.com/mic")
    merge_and_write([r1, r2])

    loaded = load_spreadsheet_resources(tmp_path / RESEARCH_FILENAME)
    assert len(loaded) == 2
    urls = {r.url for r in loaded}
    assert "https://example.com/beths" in urls
    titles = {r.title for r in loaded}
    assert "The Beths @ The Tivoli, Brisbane" in titles


# ── Deduplication (Task 13) ───────────────────────────────────────────────────

def test_exact_url_duplicate_skipped(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr("agent.config.OUTPUT_DIR", tmp_path)
    r = _make_resource("Band A @ Venue X, Gold Coast", "https://example.com/a")
    merge_and_write([r])
    added, skipped, _ = merge_and_write([r])
    assert added == 0
    assert skipped == 1
    # No source should be added for same-domain same-URL
    sources = _sources_from_sheet(tmp_path / RESEARCH_FILENAME)
    assert all(s == "" for s in sources)


def test_semantic_duplicate_adds_source_different_domain(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """Same act+date from a different website → new row skipped, URL added to Sources."""
    monkeypatch.setattr("agent.config.OUTPUT_DIR", tmp_path)
    d = (date.today() + timedelta(days=7)).isoformat()
    r1 = Resource(title="The Beths @ The Tivoli, Brisbane", url="https://ticketek.com.au/beths", date=d)
    r2 = Resource(title="The Beths @ The Tivoli, Brisbane", url="https://oztix.com.au/beths", date=d)

    merge_and_write([r1])
    added, skipped, _ = merge_and_write([r2])

    assert added == 0
    assert skipped == 1
    urls = _urls_from_sheet(tmp_path / RESEARCH_FILENAME)
    assert len([u for u in urls if u.startswith("http")]) == 1
    sources = _sources_from_sheet(tmp_path / RESEARCH_FILENAME)
    assert any("oztix.com.au" in s for s in sources)


def test_semantic_duplicate_venue_variation_ignored(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """Same act+date with slightly different venue text → still treated as duplicate."""
    monkeypatch.setattr("agent.config.OUTPUT_DIR", tmp_path)
    d = (date.today() + timedelta(days=7)).isoformat()
    r1 = Resource(title="The Beths @ The Tivoli, Brisbane", url="https://ticketek.com.au/beths", date=d)
    # Venue text differs slightly ("Tivoli" vs "The Tivoli Theatre") — same gig.
    r2 = Resource(title="The Beths @ Tivoli Theatre, Brisbane", url="https://oztix.com.au/beths", date=d)

    merge_and_write([r1])
    added, skipped, _ = merge_and_write([r2])

    assert added == 0
    assert skipped == 1
    sources = _sources_from_sheet(tmp_path / RESEARCH_FILENAME)
    assert any("oztix.com.au" in s for s in sources)


def test_semantic_duplicate_same_domain_no_source(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """Same act+date from the SAME domain → source not added (same domain rule)."""
    monkeypatch.setattr("agent.config.OUTPUT_DIR", tmp_path)
    d = (date.today() + timedelta(days=7)).isoformat()
    r1 = Resource(title="The Beths @ The Tivoli, Brisbane", url="https://ticketek.com.au/beths/1", date=d)
    r2 = Resource(title="The Beths @ The Tivoli, Brisbane", url="https://ticketek.com.au/beths/2", date=d)

    merge_and_write([r1])
    merge_and_write([r2])

    sources = _sources_from_sheet(tmp_path / RESEARCH_FILENAME)
    assert all("ticketek" not in s for s in sources)


def test_partial_act_name_same_venue_date_is_duplicate(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """'Singer 1' and 'Singer 1, with Singer 2' at same venue+date → one row."""
    monkeypatch.setattr("agent.config.OUTPUT_DIR", tmp_path)
    d = (date.today() + timedelta(days=7)).isoformat()
    r1 = Resource(title="The Beths @ The Tivoli, Brisbane", url="https://ticketek.com.au/beths", date=d)
    r2 = Resource(title="The Beths, with Wax Chattels @ The Tivoli, Brisbane", url="https://oztix.com.au/beths", date=d)

    merge_and_write([r1])
    added, skipped, _ = merge_and_write([r2])

    assert added == 0
    assert skipped == 1
    urls = _urls_from_sheet(tmp_path / RESEARCH_FILENAME)
    assert len([u for u in urls if u.startswith("http")]) == 1


def test_longer_act_name_becomes_canonical(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """When partial match found, the longer act name replaces the shorter one."""
    monkeypatch.setattr("agent.config.OUTPUT_DIR", tmp_path)
    d = (date.today() + timedelta(days=7)).isoformat()
    r1 = Resource(title="The Beths @ The Tivoli, Brisbane", url="https://ticketek.com.au/beths", date=d)
    r2 = Resource(title="The Beths, with Wax Chattels @ The Tivoli, Brisbane", url="https://oztix.com.au/beths", date=d)

    merge_and_write([r1])
    merge_and_write([r2])

    wb = load_workbook(tmp_path / RESEARCH_FILENAME)
    ws = wb.active
    row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    assert row[0] == "The Beths, with Wax Chattels"


def test_partial_act_name_different_venue_not_duplicate(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """Partial act-name match but different venue → two separate rows."""
    monkeypatch.setattr("agent.config.OUTPUT_DIR", tmp_path)
    d = (date.today() + timedelta(days=7)).isoformat()
    r1 = Resource(title="The Beths @ The Tivoli, Brisbane", url="https://ticketek.com.au/beths", date=d)
    r2 = Resource(title="The Beths, with Wax Chattels @ Fortitude Music Hall, Brisbane", url="https://oztix.com.au/beths", date=d)

    added1, _, _ = merge_and_write([r1])
    added2, _, _ = merge_and_write([r2])

    assert added1 == 1
    assert added2 == 1
    urls = _urls_from_sheet(tmp_path / RESEARCH_FILENAME)
    assert len([u for u in urls if u.startswith("http")]) == 2


def test_different_date_not_a_duplicate(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """Same act but different date → two separate rows."""
    monkeypatch.setattr("agent.config.OUTPUT_DIR", tmp_path)
    d1 = (date.today() + timedelta(days=5)).isoformat()
    d2 = (date.today() + timedelta(days=12)).isoformat()
    r1 = Resource(title="The Beths @ The Tivoli, Brisbane", url="https://ticketek.com.au/beths/1", date=d1)
    r2 = Resource(title="The Beths @ The Tivoli, Brisbane", url="https://ticketek.com.au/beths/2", date=d2)

    added1, _, _ = merge_and_write([r1])
    added2, _, _ = merge_and_write([r2])

    assert added1 == 1
    assert added2 == 1
    urls = _urls_from_sheet(tmp_path / RESEARCH_FILENAME)
    assert len([u for u in urls if u.startswith("http")]) == 2
